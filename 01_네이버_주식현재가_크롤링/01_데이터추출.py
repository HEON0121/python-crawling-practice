import requests
from bs4 import BeautifulSoup
import openpyxl


fpath = r'C:\WebCrawling\01_네이버_주식현재가_크롤링\data.xlsx'
    # print(price)

# 엑셀 만들기 
wb = openpyxl.Workbook()
# 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)
# 현재 활성화된 시트 선택
ws = wb.active

# 종목 코드 리스트
codes = [
    '005930', # 삼성전자
    '000660', # sk하이닉스
    '035720'  # 카카오 
]

# 데이터 추가하기 
ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가손익'
ws['G1'] = '수익률'
row = 2
for code in codes:
    url = f'https://finance.naver.com/item/sise.naver?code={code}'
    respponse = requests.get(url)
    html = respponse.text
    soup = BeautifulSoup(html, 'html.parser')    
    price = soup.select_one('#_nowVal').text.replace(',','')
    ws[f'B{row}'] = int(price)
    row = row + 1
# 엑셀 저장하기 
wb.save(fpath)

